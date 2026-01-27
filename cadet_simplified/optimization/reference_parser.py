"""Reference data parsing from Excel sheets.

Parses experimental chromatogram data from sheets with 'Ref_' prefix
in the Excel template.

Sheet naming convention: Ref_<experiment_name>
  - Removes 'Ref_' prefix to get experiment name
  - Matches to experiment configurations by name

Column format:
  - 'time' column (required): time in seconds
  - Component columns: one column per component with concentration values
  - Column names should match component names (case-insensitive matching)
"""

from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

from .reference_data import ReferenceDataConfig


def parse_reference_sheets(
    excel_path: Path | str,
    component_names: list[str],
) -> tuple[dict[str, ReferenceDataConfig], list[str], list[str]]:
    """Parse reference data sheets from Excel file.
    
    Looks for sheets with 'Ref_' prefix and parses them as reference
    chromatogram data.
    
    Parameters
    ----------
    excel_path : Path or str
        Path to Excel file
    component_names : list[str]
        Expected component names for matching columns
        
    Returns
    -------
    tuple[dict[str, ReferenceDataConfig], list[str], list[str]]
        (reference_data dict, errors list, warnings list)
        reference_data is keyed by experiment name
    """
    excel_path = Path(excel_path)
    errors: list[str] = []
    warnings: list[str] = []
    reference_data: dict[str, ReferenceDataConfig] = {}
    
    # Build case-insensitive component lookup
    component_lookup = {name.lower(): name for name in component_names}
    
    try:
        xl = pd.ExcelFile(excel_path)
    except Exception as e:
        errors.append(f"Failed to open Excel file: {e}")
        return reference_data, errors, warnings
    
    # Find reference sheets
    ref_sheets = [s for s in xl.sheet_names if s.startswith("Ref_")]
    
    if not ref_sheets:
        # Not an error - reference data is optional
        return reference_data, errors, warnings
    
    for sheet_name in ref_sheets:
        # Extract experiment name from sheet name
        experiment_name = sheet_name[4:]  # Remove 'Ref_' prefix
        
        if not experiment_name:
            errors.append(f"Sheet '{sheet_name}' has no experiment name after 'Ref_'")
            continue
        
        try:
            df = pd.read_excel(xl, sheet_name=sheet_name)
        except Exception as e:
            errors.append(f"Failed to read sheet '{sheet_name}': {e}")
            continue
        
        if df.empty:
            warnings.append(f"Sheet '{sheet_name}' is empty, skipping")
            continue
        
        # Parse the reference data from this sheet
        ref_config, sheet_errors, sheet_warnings = _parse_reference_sheet(
            df=df,
            sheet_name=sheet_name,
            experiment_name=experiment_name,
            component_lookup=component_lookup,
        )
        
        errors.extend(sheet_errors)
        warnings.extend(sheet_warnings)
        
        if ref_config is not None:
            reference_data[experiment_name] = ref_config
    
    return reference_data, errors, warnings


def _parse_reference_sheet(
    df: pd.DataFrame,
    sheet_name: str,
    experiment_name: str,
    component_lookup: dict[str, str],
) -> tuple[ReferenceDataConfig | None, list[str], list[str]]:
    """Parse a single reference data sheet.
    
    Parameters
    ----------
    df : pd.DataFrame
        Sheet data
    sheet_name : str
        Original sheet name (for error messages)
    experiment_name : str
        Extracted experiment name
    component_lookup : dict[str, str]
        Lowercase -> original component name mapping
        
    Returns
    -------
    tuple[ReferenceDataConfig | None, list[str], list[str]]
        (config or None, errors, warnings)
    """
    errors: list[str] = []
    warnings: list[str] = []
    
    # Normalize column names
    df.columns = [str(c).strip() for c in df.columns]
    columns_lower = {c.lower(): c for c in df.columns}
    
    # Find time column
    time_col = None
    for candidate in ['time', 'time_s', 'time (s)', 't']:
        if candidate in columns_lower:
            time_col = columns_lower[candidate]
            break
    
    if time_col is None:
        errors.append(f"Sheet '{sheet_name}': No 'time' column found")
        return None, errors, warnings
    
    # Extract time values
    time_values = df[time_col].values
    
    # Remove any NaN rows
    valid_mask = ~np.isnan(time_values)
    if not np.all(valid_mask):
        n_removed = np.sum(~valid_mask)
        warnings.append(
            f"Sheet '{sheet_name}': Removed {n_removed} rows with missing time values"
        )
        time_values = time_values[valid_mask]
    
    if len(time_values) == 0:
        errors.append(f"Sheet '{sheet_name}': No valid time values")
        return None, errors, warnings
    
    # Find component columns
    concentrations: dict[str, np.ndarray] = {}
    
    for col in df.columns:
        if col == time_col:
            continue
        
        col_lower = col.lower()
        
        # Try to match to a component name
        matched_component = None
        if col_lower in component_lookup:
            matched_component = component_lookup[col_lower]
        else:
            # Try partial matching
            for comp_lower, comp_name in component_lookup.items():
                if comp_lower in col_lower or col_lower in comp_lower:
                    matched_component = comp_name
                    if col_lower != comp_lower:
                        warnings.append(
                            f"Sheet '{sheet_name}': Column '{col}' matched to "
                            f"component '{comp_name}' (fuzzy match)"
                        )
                    break
        
        if matched_component is None:
            warnings.append(
                f"Sheet '{sheet_name}': Column '{col}' doesn't match any component, skipping"
            )
            continue
        
        # Extract concentration values
        conc_values = df[col].values[valid_mask]
        
        # Handle NaN in concentration (replace with 0)
        nan_mask = np.isnan(conc_values)
        if np.any(nan_mask):
            n_nan = np.sum(nan_mask)
            warnings.append(
                f"Sheet '{sheet_name}': {n_nan} NaN values in '{col}' replaced with 0"
            )
            conc_values = np.nan_to_num(conc_values, nan=0.0)
        
        concentrations[matched_component] = conc_values
    
    if not concentrations:
        errors.append(
            f"Sheet '{sheet_name}': No valid component columns found. "
            f"Expected columns matching: {list(component_lookup.values())}"
        )
        return None, errors, warnings
    
    # Create config
    try:
        config = ReferenceDataConfig(
            experiment_name=experiment_name,
            time=time_values,
            concentrations=concentrations,
            unit_info={"time": "s", "concentration": "mM"},
        )
    except Exception as e:
        errors.append(f"Sheet '{sheet_name}': Failed to create config: {e}")
        return None, errors, warnings
    
    return config, errors, warnings


def generate_reference_template(
    experiment_names: list[str],
    component_names: list[str],
    n_example_rows: int = 10,
) -> dict[str, pd.DataFrame]:
    """Generate template DataFrames for reference data sheets.
    
    Creates empty/example sheets for each experiment.
    
    Parameters
    ----------
    experiment_names : list[str]
        Names of experiments to create sheets for
    component_names : list[str]
        Component names for column headers
    n_example_rows : int
        Number of example rows to include
        
    Returns
    -------
    dict[str, pd.DataFrame]
        Sheet name -> DataFrame mapping
    """
    sheets = {}
    
    for exp_name in experiment_names:
        sheet_name = f"Ref_{exp_name}"
        
        # Create example data
        data = {
            "time": np.linspace(0, 3600, n_example_rows),  # 0-60 min in seconds
        }
        
        # Add component columns with placeholder values
        for comp_name in component_names:
            if comp_name.lower() == "salt":
                # Salt typically has different profile
                data[comp_name] = [np.nan] * n_example_rows
            else:
                data[comp_name] = [np.nan] * n_example_rows
        
        sheets[sheet_name] = pd.DataFrame(data)
    
    return sheets


def add_reference_sheets_to_workbook(
    workbook,
    experiment_names: list[str],
    component_names: list[str],
):
    """Add reference data sheets to an openpyxl workbook.
    
    Parameters
    ----------
    workbook : openpyxl.Workbook
        Excel workbook to add sheets to
    experiment_names : list[str]
        Experiment names
    component_names : list[str]
        Component names for columns
    """
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.comments import Comment
    
    templates = generate_reference_template(experiment_names, component_names)
    
    for sheet_name, df in templates.items():
        ws = workbook.create_sheet(title=sheet_name)
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
            for c_idx, value in enumerate(row):
                cell = ws.cell(row=r_idx + 1, column=c_idx + 1, value=value)
        
        # Add comment to time column header
        ws.cell(row=1, column=1).comment = Comment(
            "Time in seconds. Required column.",
            "cadet_simplified"
        )
        
        # Add comments to component columns
        for c_idx, comp_name in enumerate(component_names, start=2):
            ws.cell(row=1, column=c_idx).comment = Comment(
                f"Concentration of {comp_name} in mM. "
                "Leave empty or delete rows with no data.",
                "cadet_simplified"
            )
