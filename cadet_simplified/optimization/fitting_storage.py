"""Storage interfaces and implementations for fitting results.

Extends the existing file storage to handle optimization/fitting results.
"""

import json
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, TYPE_CHECKING

import pandas as pd

if TYPE_CHECKING:
    from .fitting_problem import FittingProblem
    from .fitting_result import FittingResult
    from .reference_data import ReferenceDataConfig


@dataclass
class FittingRunInfo:
    """Metadata about a stored fitting run.
    
    Attributes
    ----------
    run_id : str
        Unique identifier for this run
    name : str
        User-provided name for the fitting problem
    created_at : datetime
        When the run was saved
    n_experiments : int
        Number of experiments fitted
    n_variables : int
        Number of optimization variables
    success : bool
        Whether optimization succeeded
    final_objective : float or list
        Final objective value(s)
    runtime_seconds : float
        Total optimization runtime
    """
    run_id: str
    name: str
    created_at: datetime
    n_experiments: int
    n_variables: int
    success: bool
    final_objective: float | list | None
    runtime_seconds: float


@dataclass
class FittingStorage:
    """Storage handler for fitting/optimization results.
    
    Stores fitting results in a structured directory format:
    
        {storage_dir}/
        └── fittings/
            └── {run_id}/
                ├── config.json           # Problem configuration
                ├── result.json           # Fitted parameters and objectives
                ├── reference_data/       # Reference chromatograms
                │   ├── exp_1.parquet
                │   └── exp_2.parquet
                ├── history.parquet       # Optimization history
                └── comparison_plots/     # Optional plots
    
    Attributes
    ----------
    storage_dir : Path
        Base storage directory
        
    Example
    -------
    >>> storage = FittingStorage("./results")
    >>> run_id = storage.save(fitting_result)
    >>> loaded = storage.load(run_id)
    >>> storage.list_runs()
    """
    storage_dir: Path
    
    def __init__(self, storage_dir: Path | str):
        """Initialize fitting storage.
        
        Parameters
        ----------
        storage_dir : Path or str
            Base storage directory
        """
        self.storage_dir = Path(storage_dir)
        self._fittings_dir = self.storage_dir / "fittings"
        self._fittings_dir.mkdir(parents=True, exist_ok=True)
    
    def save(
        self,
        result: "FittingResult",
        run_id: str | None = None,
    ) -> str:
        """Save fitting result to storage.
        
        Parameters
        ----------
        result : FittingResult
            Result to save
        run_id : str, optional
            Custom run ID. If None, generates from timestamp.
            
        Returns
        -------
        str
            Run ID for the saved result
        """
        # Generate run ID if not provided
        if run_id is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            name_slug = result.problem.name.replace(" ", "_")[:20]
            run_id = f"{timestamp}_{name_slug}"
        
        run_dir = self._fittings_dir / run_id
        run_dir.mkdir(parents=True, exist_ok=True)
        
        # Save problem configuration
        config_data = self._serialize_problem(result.problem)
        config_path = run_dir / "config.json"
        with open(config_path, 'w') as f:
            json.dump(config_data, f, indent=2, default=str)
        
        # Save result
        result_data = result.to_dict()
        result_data['created_at'] = datetime.now().isoformat()
        result_path = run_dir / "result.json"
        with open(result_path, 'w') as f:
            json.dump(result_data, f, indent=2, default=str)
        
        # Save reference data
        ref_dir = run_dir / "reference_data"
        ref_dir.mkdir(exist_ok=True)
        for exp_name, ref_config in result.problem.reference_data.items():
            ref_df = ref_config.to_dataframe()
            ref_df.to_parquet(ref_dir / f"{exp_name}.parquet")
        
        # Save optimization history
        if not result.history.empty:
            history_path = run_dir / "history.parquet"
            # Handle complex types in history
            history_df = self._serialize_history(result.history)
            history_df.to_parquet(history_path)
        
        return run_id
    
    def load(self, run_id: str) -> "FittingResult":
        """Load fitting result from storage.
        
        Parameters
        ----------
        run_id : str
            Run ID to load
            
        Returns
        -------
        FittingResult
            Loaded result
            
        Raises
        ------
        FileNotFoundError
            If run ID not found
        """
        from .fitting_result import FittingResult
        from .fitting_problem import FittingProblem
        from .reference_data import ReferenceDataConfig
        
        run_dir = self._fittings_dir / run_id
        if not run_dir.exists():
            raise FileNotFoundError(f"Fitting run '{run_id}' not found")
        
        # Load result
        result_path = run_dir / "result.json"
        with open(result_path, 'r') as f:
            result_data = json.load(f)
        
        # Load config
        config_path = run_dir / "config.json"
        with open(config_path, 'r') as f:
            config_data = json.load(f)
        
        # Load history
        history_path = run_dir / "history.parquet"
        if history_path.exists():
            history = pd.read_parquet(history_path)
        else:
            history = pd.DataFrame()
        
        # Load reference data
        reference_data = {}
        ref_dir = run_dir / "reference_data"
        if ref_dir.exists():
            for ref_file in ref_dir.glob("*.parquet"):
                exp_name = ref_file.stem
                ref_df = pd.read_parquet(ref_file)
                reference_data[exp_name] = ReferenceDataConfig.from_dataframe(
                    ref_df, exp_name
                )
        
        # Reconstruct a minimal FittingResult
        # Note: Full problem reconstruction requires operation mode, etc.
        # This creates a partial result for inspection
        return self._reconstruct_result(
            result_data=result_data,
            config_data=config_data,
            history=history,
            reference_data=reference_data,
        )
    
    def list_runs(self) -> list[FittingRunInfo]:
        """List all stored fitting runs.
        
        Returns
        -------
        list[FittingRunInfo]
            Information about each stored run
        """
        runs = []
        
        for run_dir in sorted(self._fittings_dir.iterdir(), reverse=True):
            if not run_dir.is_dir():
                continue
            
            result_path = run_dir / "result.json"
            if not result_path.exists():
                continue
            
            try:
                with open(result_path, 'r') as f:
                    result_data = json.load(f)
                
                created_at_str = result_data.get('created_at')
                if created_at_str:
                    created_at = datetime.fromisoformat(created_at_str)
                else:
                    # Fallback to directory modification time
                    created_at = datetime.fromtimestamp(run_dir.stat().st_mtime)
                
                runs.append(FittingRunInfo(
                    run_id=run_dir.name,
                    name=result_data.get('problem_name', 'Unknown'),
                    created_at=created_at,
                    n_experiments=len(result_data.get('variable_info', [])),
                    n_variables=len(result_data.get('fitted_parameters', {})),
                    success=result_data.get('success', False),
                    final_objective=result_data.get('final_objective'),
                    runtime_seconds=result_data.get('runtime_seconds', 0),
                ))
            except Exception:
                # Skip malformed entries
                continue
        
        return runs
    
    def delete(self, run_id: str) -> bool:
        """Delete a fitting run.
        
        Parameters
        ----------
        run_id : str
            Run ID to delete
            
        Returns
        -------
        bool
            True if deleted, False if not found
        """
        import shutil
        
        run_dir = self._fittings_dir / run_id
        if not run_dir.exists():
            return False
        
        shutil.rmtree(run_dir)
        return True
    
    def get_run_path(self, run_id: str) -> Path:
        """Get path to a fitting run directory.
        
        Parameters
        ----------
        run_id : str
            Run ID
            
        Returns
        -------
        Path
            Path to run directory
        """
        return self._fittings_dir / run_id
    
    def _serialize_problem(self, problem: "FittingProblem") -> dict:
        """Serialize problem configuration."""
        return {
            'name': problem.name,
            'column_model': problem.column_binding.column_model,
            'binding_model': problem.column_binding.binding_model,
            'component_names': problem.component_system,
            'experiments_to_fit': problem.experiments_to_fit,
            'metric': problem.metric,
            'comparison_components': problem.comparison_components,
            'selected_variables': [
                {
                    'parameter_name': v.parameter.name,
                    'lb': v.lb,
                    'ub': v.ub,
                    'transform': v.transform,
                    'components': v.components,
                }
                for v in problem.selected_variables
            ],
        }
    
    def _serialize_history(self, history: pd.DataFrame) -> pd.DataFrame:
        """Serialize history DataFrame, handling complex types."""
        df = history.copy()
        
        # Convert complex columns to JSON strings
        for col in df.columns:
            if df[col].dtype == object:
                df[col] = df[col].apply(
                    lambda x: json.dumps(x) if isinstance(x, (list, dict, tuple)) else x
                )
        
        return df
    
    def _reconstruct_result(
        self,
        result_data: dict,
        config_data: dict,
        history: pd.DataFrame,
        reference_data: dict,
    ) -> "FittingResult":
        """Reconstruct FittingResult from stored data.
        
        Note: This creates a partial result without full problem reconstruction.
        """
        from .fitting_result import FittingResult
        
        # Create a minimal mock problem for the result
        @dataclass
        class MockProblem:
            name: str
            experiments_to_fit: list
            reference_data: dict
            
        mock_problem = MockProblem(
            name=result_data.get('problem_name', 'Unknown'),
            experiments_to_fit=config_data.get('experiments_to_fit', []),
            reference_data=reference_data,
        )
        
        return FittingResult(
            problem=mock_problem,  # type: ignore
            success=result_data.get('success', False),
            fitted_parameters=result_data.get('fitted_parameters', {}),
            final_objective=result_data.get('final_objective'),
            n_generations=result_data.get('n_generations', 0),
            runtime_seconds=result_data.get('runtime_seconds', 0),
            history=history,
            pareto_x=result_data.get('pareto_x'),
            pareto_f=result_data.get('pareto_f'),
            variable_info=result_data.get('variable_info', []),
        )


def export_fitting_result_to_excel(
    result: "FittingResult",
    output_path: Path | str,
):
    """Export fitting result to Excel file.
    
    Creates a workbook with:
    - Summary sheet with fitted parameters
    - History sheet with convergence data
    - Reference data sheets
    
    Parameters
    ----------
    result : FittingResult
        Result to export
    output_path : Path or str
        Output Excel file path
    """
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    output_path = Path(output_path)
    wb = Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    ws_summary.cell(row=1, column=1, value="Fitting Result Summary")
    ws_summary.cell(row=2, column=1, value="Problem Name")
    ws_summary.cell(row=2, column=2, value=result.problem.name)
    ws_summary.cell(row=3, column=1, value="Success")
    ws_summary.cell(row=3, column=2, value=str(result.success))
    ws_summary.cell(row=4, column=1, value="Runtime (s)")
    ws_summary.cell(row=4, column=2, value=result.runtime_seconds)
    ws_summary.cell(row=5, column=1, value="Generations")
    ws_summary.cell(row=5, column=2, value=result.n_generations)
    
    # Fitted parameters
    ws_summary.cell(row=7, column=1, value="Fitted Parameters")
    row = 8
    for name, value in result.fitted_parameters.items():
        if isinstance(value, dict):
            for comp, val in value.items():
                ws_summary.cell(row=row, column=1, value=f"{name}_{comp}")
                ws_summary.cell(row=row, column=2, value=val)
                row += 1
        else:
            ws_summary.cell(row=row, column=1, value=name)
            ws_summary.cell(row=row, column=2, value=value)
            row += 1
    
    # History sheet
    if not result.history.empty:
        ws_history = wb.create_sheet(title="History")
        for r_idx, row in enumerate(dataframe_to_rows(result.history, index=False, header=True)):
            for c_idx, value in enumerate(row):
                # Handle complex values
                if isinstance(value, (list, dict)):
                    value = str(value)
                ws_history.cell(row=r_idx + 1, column=c_idx + 1, value=value)
    
    # Reference data sheets
    for exp_name, ref_config in result.problem.reference_data.items():
        ws_ref = wb.create_sheet(title=f"Ref_{exp_name}"[:31])  # Excel sheet name limit
        ref_df = ref_config.to_dataframe()
        for r_idx, row in enumerate(dataframe_to_rows(ref_df, index=False, header=True)):
            for c_idx, value in enumerate(row):
                ws_ref.cell(row=r_idx + 1, column=c_idx + 1, value=value)
    
    wb.save(output_path)
